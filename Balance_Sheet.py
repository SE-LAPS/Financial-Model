from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

def create_balance_sheet(wb):
    """Create and format the Balance_Sheet sheet."""
    print("Creating Balance_Sheet sheet...")
    
    # Get the Balance_Sheet sheet
    ws = wb["Balance_Sheet"]
    
    # Set the title
    ws['A1'] = "BALANCE SHEET"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Column headers
    ws['A3'] = "Line Item"
    
    # Year headers
    years = ["2025", "2026", "2027", "2028", "2029"]
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)  # Start from column B
        ws[f'{col}3'] = year
        ws[f'{col}3'].font = Font(bold=True)
    
    # Make header row bold
    ws['A3'].font = Font(bold=True)
    
    # Assets section
    ws['A5'] = "ASSETS"
    ws['A5'].font = Font(bold=True)
    
    # Assets line items
    asset_items = [
        "Cash and Cash Equivalents",
        "Accounts Receivable",
        "Inventory",
        "Property, Plant & Equipment",
        "Total Assets"
    ]
    
    for i, item in enumerate(asset_items, 6):
        row = i
        ws[f'A{row}'] = item
        if item == "Total Assets":
            ws[f'A{row}'].font = Font(bold=True)
    
    # Cash and Cash Equivalents
    ws['B6'] = 500000  # Starting cash for 2025
    
    # Cash for future years will be linked to Cash_Flow in later steps
    # We'll update these formulas after creating the Cash_Flow sheet
    
    # Accounts Receivable (15% of Revenue)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}7'] = f"=Revenue_Forecast!{col}9*0.15"
    
    # Inventory (10% of COGS)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}8'] = f"=COGS_Budget!{col}9*0.1"
    
    # Property, Plant & Equipment
    ws['B9'] = 2000000  # Initial PP&E for 2025
    
    # PP&E for future years (+ CapEx - D&A)
    capex_values = [200000, 210000, 220500, 231525]
    
    for j, (year, capex) in enumerate(zip(years[1:], capex_values), 3):
        col = get_column_letter(j)
        prev_col = get_column_letter(j-1)
        ws[f'{col}9'] = f"={prev_col}9+{capex}-OPEX_Budget!{col}7"
    
    # Total Assets
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}10'] = f"=SUM({col}6:{col}9)"
        ws[f'{col}10'].font = Font(bold=True)
    
    # Liabilities and Equity section
    ws['A12'] = "LIABILITIES AND EQUITY"
    ws['A12'].font = Font(bold=True)
    
    # Liabilities line items
    liability_items = [
        "Accounts Payable",
        "Long-term Debt",
        "Total Liabilities"
    ]
    
    for i, item in enumerate(liability_items, 13):
        row = i
        ws[f'A{row}'] = item
        if item == "Total Liabilities":
            ws[f'A{row}'].font = Font(bold=True)
    
    # Accounts Payable (10% of COGS)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}13'] = f"=COGS_Budget!{col}9*0.1"
    
    # Long-term Debt
    ws['B14'] = 1000000  # Initial debt for 2025
    
    # Debt for future years (assuming 100k debt repayment each year)
    for j, year in enumerate(years[1:], 3):
        col = get_column_letter(j)
        prev_col = get_column_letter(j-1)
        ws[f'{col}14'] = f"={prev_col}14-100000"
    
    # Total Liabilities
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}15'] = f"={col}13+{col}14"
        ws[f'{col}15'].font = Font(bold=True)
    
    # Equity section
    ws['A17'] = "Equity"
    ws['A17'].font = Font(bold=True)
    
    # Equity line items
    equity_items = [
        "Common Stock",
        "Retained Earnings",
        "Total Equity"
    ]
    
    for i, item in enumerate(equity_items, 18):
        row = i
        ws[f'A{row}'] = item
        if item == "Total Equity":
            ws[f'A{row}'].font = Font(bold=True)
    
    # Common Stock
    ws['B18'] = 1000000  # Initial common stock for 2025
    
    # Common stock for future years (remains constant)
    for j, year in enumerate(years[1:], 3):
        col = get_column_letter(j)
        prev_col = get_column_letter(j-1)
        ws[f'{col}18'] = f"={prev_col}18"
    
    # Retained Earnings
    ws['B19'] = 500000  # Initial retained earnings for 2025
    
    # Retained earnings for future years (+ Net Income - Dividends)
    dividend_values = [50000, 55000, 60500, 66550]
    
    for j, (year, dividend) in enumerate(zip(years[1:], dividend_values), 3):
        col = get_column_letter(j)
        prev_col = get_column_letter(j-1)
        ws[f'{col}19'] = f"={prev_col}19+Income_Statement!{prev_col}14-{dividend}"
    
    # Total Equity
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}20'] = f"={col}18+{col}19"
        ws[f'{col}20'].font = Font(bold=True)
    
    # Total Liabilities and Equity
    ws['A21'] = "Total Liabilities and Equity"
    ws['A21'].font = Font(bold=True)
    
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}21'] = f"={col}15+{col}20"
        ws[f'{col}21'].font = Font(bold=True)
    
    # Balance Check (should be zero)
    ws['A23'] = "Balance Check (Assets - Liabilities - Equity)"
    
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}23'] = f"={col}10-{col}21"
        
        # Add conditional formatting to highlight non-zero values in red
        red_font = Font(color="FF0000")  # Red font
        rule = CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=False, font=red_font)
        ws.conditional_formatting.add(f'{col}23', rule)
    
    # Format all monetary values with thousands separator
    for row in range(6, 24):
        if row != 17:  # Skip the "Equity" header row
            for col in range(2, 7):  # Columns B to F
                cell = ws[f'{get_column_letter(col)}{row}']
                cell.number_format = '#,##0'
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 7):  # Columns B to F
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add background colors to key rows
    asset_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')      # Light blue
    liability_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # Light orange
    equity_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')     # Light green
    
    # Apply fills
    for col in range(1, 7):  # Columns A to F
        col_letter = get_column_letter(col)
        ws[f'{col_letter}10'].fill = asset_fill          # Total Assets
        ws[f'{col_letter}15'].fill = liability_fill      # Total Liabilities
        ws[f'{col_letter}20'].fill = equity_fill         # Total Equity
        ws[f'{col_letter}21'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Total L&E
    
    print("Balance_Sheet sheet created successfully")