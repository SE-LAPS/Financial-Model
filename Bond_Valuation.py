from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_bond_valuation(wb):
    """Create and format the Bond_Valuation sheet."""
    print("Creating Bond_Valuation sheet...")
    
    # Get the Bond_Valuation sheet
    ws = wb["Bond_Valuation"]
    
    # Set the title
    ws['A1'] = "BOND VALUATION MODEL"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Bond Parameters section
    ws['A3'] = "Bond Parameters"
    ws['A3'].font = Font(bold=True)
    
    # Bond parameters
    params = [
        ("Par Value", 1000),
        ("Coupon Rate", 0.05),
        ("Years to Maturity", 10),
        ("Payments per Year", 2),
        ("Required Yield", 0.06)
    ]
    
    # Add data validation
    dv_positive = DataValidation(type="decimal", operator="greaterThan", formula1="0")
    dv_positive.error = "Value must be greater than 0"
    dv_positive.errorTitle = "Invalid Input"
    
    dv_rate = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
    dv_rate.error = "Rate must be between 0 and 1"
    dv_rate.errorTitle = "Invalid Rate"
    
    ws.add_data_validation(dv_positive)
    ws.add_data_validation(dv_rate)
    
    for i, (param, value) in enumerate(params, 4):
        row = i
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        
        # Format percentages and currency
        if "Rate" in param or "Yield" in param:
            ws[f'B{row}'].number_format = '0.00%'
            dv_rate.add(f'B{row}')
        elif "Value" in param:
            ws[f'B{row}'].number_format = '$#,##0'
            dv_positive.add(f'B{row}')
        else:
            dv_positive.add(f'B{row}')
    
    # Calculations
    ws['A10'] = "Total Periods"
    ws['B10'] = "=B6*B7"
    
    ws['A11'] = "Periodic Coupon Payment"
    ws['B11'] = "=B4*B5/B7"
    ws['B11'].number_format = '$#,##0.00'
    
    ws['A12'] = "Periodic Yield Rate"
    ws['B12'] = "=B8/B7"
    ws['B12'].number_format = '0.00%'
    
    # Bond Cash Flow Table
    ws['A14'] = "Period"
    ws['B14'] = "Cash Flow"
    ws['C14'] = "Present Value"
    ws['A14'].font = Font(bold=True)
    ws['B14'].font = Font(bold=True)
    ws['C14'].font = Font(bold=True)
    
    # Create full cash flow table
    for i in range(1, 21):  # Assuming max 10 years with semi-annual payments
        row = i + 14
        ws[f'A{row}'] = i
        
        # Cash flow formula
        if i == 20:  # Last period
            ws[f'B{row}'] = "=B11+B4"  # Final coupon + par value
        else:
            ws[f'B{row}'] = "=B11"  # Regular coupon payment
        
        # Present value formula
        ws[f'C{row}'] = f"=B{row}/(1+B12)^A{row}"
        
        # Format with currency
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].number_format = '$#,##0.00'
    
    # Bond Value
    ws['A35'] = "Bond Value"
    ws['A35'].font = Font(bold=True)
    ws['B35'] = "=SUM(C15:C34)"  # Sum of all PV cash flows
    ws['B35'].number_format = '$#,##0.00'
    ws['B35'].font = Font(bold=True)
    
    # YTM Calculation
    ws['A37'] = "Bond YTM Calculation"
    ws['A37'].font = Font(bold=True)
    
    ws['A38'] = "Current Bond Price"
    ws['B38'] = 950
    ws['B38'].number_format = '$#,##0.00'
    dv_positive.add('B38')
    
    ws['A39'] = "Approximate YTM"
    ws['B39'] = "=((B11*B7)+((B4-B38)/B6))/((B4+B38)/2)"
    ws['B39'].number_format = '0.00%'
    
    # Add error checking
    ws['A41'] = "Validation Checks"
    ws['A41'].font = Font(bold=True)
    
    ws['A42'] = "Price/Par Value Ratio"
    ws['B42'] = "=B35/B4"
    ws['B42'].number_format = '0.00%'
    
    ws['A43'] = "Status"
    ws['B43'] = '=IF(AND(B35>0,B39>0),"Valid","Check Inputs")'
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    print("Bond_Valuation sheet created successfully")