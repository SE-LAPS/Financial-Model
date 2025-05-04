from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference

def create_capital_budgeting(wb):
    """Create and format the Capital_Budgeting sheet."""
    print("Creating Capital_Budgeting sheet...")
    
    # Get the Capital_Budgeting sheet
    ws = wb["Capital_Budgeting"]
    
    # Set the title
    ws['A1'] = "CAPITAL BUDGETING MODEL"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Project Parameters section
    ws['A3'] = "Project Parameters"
    ws['A3'].font = Font(bold=True)
    
    # Project parameters
    params = [
        ("Initial Investment", 500000),
        ("Project Life (Years)", 5),
        ("Discount Rate", "=Assumptions!B22"),
        ("Salvage Value", 50000)
    ]
    
    # Add data validation
    dv_positive = DataValidation(type="decimal", operator="greaterThan", formula1="0")
    dv_positive.error = "Value must be greater than 0"
    dv_positive.errorTitle = "Invalid Input"
    
    dv_rate = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
    dv_rate.error = "Rate must be between 0 and 1"
    dv_rate.errorTitle = "Invalid Rate"
    
    ws.add_data_validation(dv_positive)
    ws.add_data_validation(dv_rate)
    
    for i, (param, value) in enumerate(params, 4):
        row = i
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        
        # Format percentages and currency
        if param == "Discount Rate":
            ws[f'B{row}'].number_format = '0.00%'
            dv_rate.add(f'B{row}')
        elif "Value" in param or "Investment" in param:
            ws[f'B{row}'].number_format = '#,##0'
            dv_positive.add(f'B{row}')
        else:
            dv_positive.add(f'B{row}')
    
    # Project Cash Flows section
    ws['A9'] = "Year"
    ws['B9'] = "Cash Flow"
    ws['C9'] = "Discounted Cash Flow"
    ws['D9'] = "Cumulative Cash Flow"
    ws['E9'] = "Cumulative Discounted Cash Flow"
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}9'].font = Font(bold=True)
    
    # Cash flow by year
    cash_flows = [
        (0, "=-B4"),  # Initial investment (negative)
        (1, 120000),
        (2, 150000),
        (3, 180000),
        (4, 200000),
        (5, "=220000+B7")  # Final year cash flow + salvage value
    ]
    
    for i, (year, cf) in enumerate(cash_flows, 10):
        row = i
        ws[f'A{row}'] = year
        ws[f'B{row}'] = cf
        
        # Discounted cash flow
        if isinstance(cf, str):
            ws[f'C{row}'] = f"={cf}/(1+$B$6)^A{row}"
        else:
            ws[f'C{row}'] = f"=B{row}/(1+$B$6)^A{row}"
        
        # Cumulative cash flow
        if year == 0:
            ws[f'D{row}'] = f"=B{row}"
            ws[f'E{row}'] = f"=C{row}"
        else:
            ws[f'D{row}'] = f"=D{row-1}+B{row}"
            ws[f'E{row}'] = f"=E{row-1}+C{row}"
        
        # Format numbers
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].number_format = '#,##0'
    
    # NPV Calculation section
    ws['A17'] = "NPV Calculation"
    ws['A17'].font = Font(bold=True)
    
    ws['A18'] = "Present Value of Cash Flows"
    ws['B18'] = "=E15"  # Sum of all discounted cash flows
    ws['B18'].number_format = '#,##0'
    
    ws['A19'] = "NPV Decision"
    ws['B19'] = '=IF(B18>0,"Accept Project","Reject Project")'
    ws['B19'].font = Font(bold=True)
    
    # IRR Calculation section
    ws['A21'] = "IRR Calculation"
    ws['A21'].font = Font(bold=True)
    
    ws['A22'] = "Internal Rate of Return (IRR)"
    ws['B22'] = "=IRR(B10:B15)"
    ws['B22'].number_format = '0.00%'
    
    ws['A23'] = "IRR Decision"
    ws['B23'] = '=IF(B22>B6,"Accept Project","Reject Project")'
    ws['B23'].font = Font(bold=True)
    
    # Payback Period Calculation section
    ws['A25'] = "Payback Period Calculation"
    ws['A25'].font = Font(bold=True)
    
    ws['A26'] = "Payback Period (Years)"
    ws['B26'] = '=MATCH(0,D10:D15,1)-1+ABS(INDEX(D10:D15,MATCH(0,D10:D15,1)-1))/INDEX(B10:B15,MATCH(0,D10:D15,1))'
    ws['B26'].number_format = '0.00'
    
    ws['A27'] = "Discounted Payback Period (Years)"
    ws['B27'] = '=MATCH(0,E10:E15,1)-1+ABS(INDEX(E10:E15,MATCH(0,E10:E15,1)-1))/INDEX(C10:C15,MATCH(0,E10:E15,1))'
    ws['B27'].number_format = '0.00'
    
    # Create a line chart for cumulative cash flows
    chart = LineChart()
    chart.title = "Cumulative Cash Flows"
    chart.style = 10
    chart.x_axis.title = "Year"
    chart.y_axis.title = "Cash Flow"
    
    # Add data series
    cats = Reference(ws, min_col=1, min_row=10, max_row=15)
    values1 = Reference(ws, min_col=4, min_row=9, max_row=15)
    values2 = Reference(ws, min_col=5, min_row=9, max_row=15)
    
    chart.add_data(values1, titles_from_data=True)
    chart.add_data(values2, titles_from_data=True)
    chart.set_categories(cats)
    
    # Add the chart to the worksheet
    ws.add_chart(chart, "A30")
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20
    
    print("Capital_Budgeting sheet created successfully")