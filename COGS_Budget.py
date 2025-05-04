from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def create_cogs_budget(wb):
    """Create and format the COGS_Budget sheet."""
    print("Creating COGS_Budget sheet...")
    
    # Get the COGS_Budget sheet
    ws = wb["COGS_Budget"]
    
    # Set the title
    ws['A1'] = "COST OF GOODS SOLD BUDGET"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Column headers
    ws['A3'] = "Category"
    
    # Year headers
    years = ["2025", "2026", "2027", "2028", "2029"]
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)  # Start from column B
        ws[f'{col}3'] = year
        ws[f'{col}3'].font = Font(bold=True)
    
    # Make header row bold
    ws['A3'].font = Font(bold=True)
    
    # COGS for each product line
    product_lines = ["COGS Product Line 1", "COGS Product Line 2", "COGS Product Line 3"]
    
    for i, product in enumerate(product_lines, 5):
        row = i
        ws[f'A{row}'] = product
        
        # COGS formulas for each year
        revenue_row = i - 4 + 5  # Map to the corresponding row in Revenue_Forecast (5, 6, 7)
        
        for j, year in enumerate(years, 2):  # Start from column B
            col = get_column_letter(j)
            ws[f'{col}{row}'] = f"=Revenue_Forecast!{col}{revenue_row}*Assumptions!B17"
    
    # Total COGS row
    ws['A9'] = "Total COGS"
    ws['A9'].font = Font(bold=True)
    
    # Sum formulas for total COGS
    for i, year in enumerate(years, 2):  # Start from column B
        col = get_column_letter(i)
        ws[f'{col}9'] = f"=SUM({col}5:{col}7)"
        ws[f'{col}9'].font = Font(bold=True)
    
    # Format all COGS values with thousands separator
    for row in range(5, 10):
        for col in range(2, 7):  # Columns B to F
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.number_format = '#,##0'
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    for col in range(2, 7):  # Columns B to F
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add a light red background to the total row
    total_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    for col in range(1, 7):  # Columns A to F
        ws[f'{get_column_letter(col)}9'].fill = total_fill
    
    print("COGS_Budget sheet created successfully")