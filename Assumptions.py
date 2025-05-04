from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def create_assumptions(wb):
    """Create and format the Assumptions sheet."""
    print("Creating Assumptions sheet...")
    
    # Get the Assumptions sheet
    ws = wb["Assumptions"]
    
    # Set the title
    ws['A1'] = "MODEL ASSUMPTIONS"
    ws['A1'].font = Font(bold=True, size=14)
    
    # General Assumptions section
    ws['A3'] = "General Assumptions"
    ws['A3'].font = Font(bold=True)
    
    # General assumptions data
    assumptions = [
        ("Base Year", 2025),
        ("Forecast Period (Years)", 5),
        ("Tax Rate", 0.24),
        ("Inflation Rate", 0.03)
    ]
    
    for i, (label, value) in enumerate(assumptions, 4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        
        # Format percentages
        if "Rate" in label and value < 1:
            ws[f'B{i}'].number_format = '0.00%'
    
    # Revenue Growth Assumptions section
    ws['A10'] = "Revenue Growth Assumptions"
    ws['A10'].font = Font(bold=True)
    
    # Revenue growth data
    growth_assumptions = [
        ("Product Line 1 Growth", 0.05),
        ("Product Line 2 Growth", 0.07),
        ("Product Line 3 Growth", 0.04)
    ]
    
    for i, (label, value) in enumerate(growth_assumptions, 11):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'B{i}'].number_format = '0.00%'
    
    # Cost Assumptions section
    ws['A16'] = "Cost Assumptions"
    ws['A16'].font = Font(bold=True)
    
    # Cost assumptions data
    cost_assumptions = [
        ("COGS as % of Revenue", 0.6),
        ("SG&A as % of Revenue", 0.15),
        ("R&D as % of Revenue", 0.08)
    ]
    
    for i, (label, value) in enumerate(cost_assumptions, 17):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'B{i}'].number_format = '0.00%'
    
    # Discount Rate
    ws['A22'] = "Discount Rate (WACC)"
    ws['B22'] = 0.1
    ws['B22'].number_format = '0.00%'
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    # Add a light yellow background to the section headers
    section_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    for row in [3, 10, 16]:
        ws[f'A{row}'].fill = section_fill
    
    print("Assumptions sheet created successfully")