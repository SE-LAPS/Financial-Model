from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def create_contributions(wb):
    """Create and format the Contributions sheet."""
    print("Creating Contributions sheet...")
    
    # Get the Contributions sheet
    ws = wb["Contributions"]
    
    # Set the title
    ws['A1'] = "TEAM MEMBER CONTRIBUTIONS"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Column headers
    headers = ["Team Member Name", "Student ID", "Contribution Description", "Contribution %"]
    
    for i, header in enumerate(headers, 1):
        col = get_column_letter(i)
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True)
    
    # Example team members and contributions
    team_members = [
        ("Team Member 1", "ID12345", "Dashboard, Business Overview, Assumptions", 10),
        ("Team Member 2", "ID23456", "Revenue Forecast, COGS Budget", 10),
        ("Team Member 3", "ID34567", "OPEX Budget, Income Statement", 10),
        ("Team Member 4", "ID45678", "Balance Sheet, Cash Flow", 10),
        ("Team Member 5", "ID56789", "Stock Valuation", 10),
        ("Team Member 6", "ID67890", "Bond Valuation", 10),
        ("Team Member 7", "ID78901", "Capital Budgeting", 10),
        ("Team Member 8", "ID89012", "Sensitivity Analysis", 10),
        ("Team Member 9", "ID90123", "Presentations, Documentation", 10),
        ("Team Member 10", "ID01234", "Quality Control, Integration", 10)
    ]
    
    for i, (name, id, desc, contrib) in enumerate(team_members, 4):
        row = i
        ws[f'A{row}'] = name
        ws[f'B{row}'] = id
        ws[f'C{row}'] = desc
        ws[f'D{row}'] = contrib
        ws[f'D{row}'].number_format = '0%'
    
    # Total contribution
    ws['A15'] = "Total"
    ws['A15'].font = Font(bold=True)
    ws['D15'] = "=SUM(D4:D13)"
    ws['D15'].number_format = '0%'
    ws['D15'].font = Font(bold=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    
    # Add background colors
    header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')  # Light blue
    total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')   # Light green
    
    # Apply fills
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}3'].fill = header_fill
    
    ws['A15'].fill = total_fill
    ws['D15'].fill = total_fill
    
    # Add alternating row colors for better readability
    for row in range(4, 14):
        if row % 2 == 0:  # Even rows
            fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light gray
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = fill
    
    # Add borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in range(3, 16):
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = thin_border
    
    print("Contributions sheet created successfully")