import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# Import all worksheet creation modules
from Dashboard import create_dashboard
from Business_Overview import create_business_overview
from Assumptions import create_assumptions
from Revenue_Forecast import create_revenue_forecast
from COGS_Budget import create_cogs_budget
from OPEX_Budget import create_opex_budget
from Income_Statement import create_income_statement
from Balance_Sheet import create_balance_sheet
from Cash_Flow import create_cash_flow
from Stock_Valuation import create_stock_valuation
from Bond_Valuation import create_bond_valuation
from Capital_Budgeting import create_capital_budgeting
from Sensitivity_Analysis import create_sensitivity_analysis
from Contributions import create_contributions

def create_financial_model(filename="Financial_Model.xlsx"):
    """Create a complete financial model Excel workbook."""
    print("Creating Financial Model...")

    # Create a new workbook
    wb = Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create all sheets
    sheets = [
        "Dashboard", "Business_Overview", "Assumptions", "Revenue_Forecast",
        "COGS_Budget", "OPEX_Budget", "Income_Statement", "Balance_Sheet",
        "Cash_Flow", "Stock_Valuation", "Bond_Valuation", "Capital_Budgeting",
        "Sensitivity_Analysis", "Contributions"
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(sheet_name)
    
    # Set up each sheet with the corresponding function
    create_dashboard(wb)
    create_business_overview(wb)
    create_assumptions(wb)
    create_revenue_forecast(wb)
    create_cogs_budget(wb)
    create_opex_budget(wb)
    create_income_statement(wb)
    create_balance_sheet(wb)
    create_cash_flow(wb)
    create_stock_valuation(wb)
    create_bond_valuation(wb)
    create_capital_budgeting(wb)
    create_sensitivity_analysis(wb)
    create_contributions(wb)
    
    # Save the workbook
    wb.save(filename)
    print(f"Financial model created successfully and saved as {filename}")

if __name__ == "__main__":
    create_financial_model()