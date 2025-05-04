from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def create_revenue_forecast(wb):
    """Create and format the Revenue_Forecast sheet."""
    print("Creating Revenue_Forecast sheet...")
    
    # Get the Revenue_Forecast sheet
    ws = wb["Revenue_Forecast"]
    
    # Set the title
    ws['A1'] = "REVENUE FORECAST"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Column headers
    ws['A3'] = "Category"
    
    # Year headers
    years = ["2025", "2026", "2027", "2028", "2029"]
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)  # Start from column B
        ws[f'{col}3'] = year
        ws[f'{col}3'].font = Font(bold=True)
    
    # Make header row bold
    ws['A3'].font = Font(bold=True)
    
    # Product lines
    product_lines = ["Product Line 1", "Product Line 2", "Product Line 3"]
    base_revenues = [1000000, 750000, 500000]
    
    for i, (product, base_rev) in enumerate(zip(product_lines, base_revenues), 5):
        row = i
        ws[f'A{row}'] = product
        ws[f'B{row}'] = base_rev
        
        # Growth formulas for subsequent years
        growth_refs = {
            "Product Line 1": "Assumptions!B11",
            "Product Line 2": "Assumptions!B12",
            "Product Line 3": "Assumptions!B13"
        }
        
        for j, year in enumerate(years[1:], 3):  # Start from 2026 (column C)
            col = get_column_letter(j)
            prev_col = get_column_letter(j-1)
            growth_ref = growth_refs[product]
            ws[f'{col}{row}'] = f"={prev_col}{row}*(1+{growth_ref})"
    
    # Total Revenue row
    ws['A9'] = "Total Revenue"
    ws['A9'].font = Font(bold=True)
    
    # Sum formulas for total revenue
    for i, year in enumerate(years, 2):  # Start from column B
        col = get_column_letter(i)
        ws[f'{col}9'] = f"=SUM({col}5:{col}7)"
        ws[f'{col}9'].font = Font(bold=True)
    
    # Format all revenue values with thousands separator
    for row in range(5, 10):
        for col in range(2, 7):  # Columns B to F
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.number_format = '#,##0'
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    for col in range(2, 7):  # Columns B to F
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add a light green background to the total row
    total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    for col in range(1, 7):  # Columns A to F
        ws[f'{get_column_letter(col)}9'].fill = total_fill
    
    print("Revenue_Forecast sheet created successfully")