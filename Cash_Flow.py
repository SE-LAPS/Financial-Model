from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def create_cash_flow(wb):
    """Create and format the Cash_Flow sheet."""
    print("Creating Cash_Flow sheet...")
    
    # Get the Cash_Flow sheet
    ws = wb["Cash_Flow"]
    
    # Set the title
    ws['A1'] = "CASH FLOW STATEMENT"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Column headers
    ws['A3'] = "Line Item"
    
    # Year headers
    years = ["2025", "2026", "2027", "2028", "2029"]
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)  # Start from column B
        ws[f'{col}3'] = year
        ws[f'{col}3'].font = Font(bold=True)
    
    # Make header row bold
    ws['A3'].font = Font(bold=True)
    
    # Operating Activities section
    ws['A5'] = "OPERATING ACTIVITIES"
    ws['A5'].font = Font(bold=True)
    
    # Operating activities line items
    op_items = [
        "Net Income",
        "Depreciation & Amortization",
        "Changes in Working Capital",
        "Cash Flow from Operating Activities"
    ]
    
    for i, item in enumerate(op_items, 6):
        row = i
        ws[f'A{row}'] = item
        if item == "Cash Flow from Operating Activities":
            ws[f'A{row}'].font = Font(bold=True)
    
    # Net Income formulas
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}6'] = f"=Income_Statement!{col}14"
    
    # Depreciation & Amortization formulas
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}7'] = f"=OPEX_Budget!{col}7"
    
    # Changes in Working Capital
    ws['B8'] = 0  # No change in WC for base year
    
    # Placeholder for Changes in Working Capital - will need complex formulas
    # For simplicity, we'll use fixed values for now
    wc_changes = [-20000, -25000, -30000, -35000]
    
    for j, (year, wc) in enumerate(zip(years[1:], wc_changes), 3):
        col = get_column_letter(j)
        ws[f'{col}8'] = wc
    
    # Cash Flow from Operating Activities
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}9'] = f"=SUM({col}6:{col}8)"
        ws[f'{col}9'].font = Font(bold=True)
    
    # Investing Activities section
    ws['A11'] = "INVESTING ACTIVITIES"
    ws['A11'].font = Font(bold=True)
    
    # Investing activities line items
    inv_items = [
        "Capital Expenditures",
        "Cash Flow from Investing Activities"
    ]
    
    for i, item in enumerate(inv_items, 12):
        row = i
        ws[f'A{row}'] = item
        if item == "Cash Flow from Investing Activities":
            ws[f'A{row}'].font = Font(bold=True)
    
    # Capital Expenditures
    capex_values = [-200000, -210000, -220500, -231525, -243101]
    
    for j, (year, capex) in enumerate(zip(years, capex_values), 2):
        col = get_column_letter(j)
        ws[f'{col}12'] = capex
    
    # Cash Flow from Investing Activities
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}13'] = f"={col}12"
        ws[f'{col}13'].font = Font(bold=True)
    
    # Financing Activities section
    ws['A15'] = "FINANCING ACTIVITIES"
    ws['A15'].font = Font(bold=True)
    
    # Financing activities line items
    fin_items = [
        "Debt Repayment",
        "Dividends Paid",
        "Cash Flow from Financing Activities"
    ]
    
    for i, item in enumerate(fin_items, 16):
        row = i
        ws[f'A{row}'] = item
        if item == "Cash Flow from Financing Activities":
            ws[f'A{row}'].font = Font(bold=True)
    
    # Debt Repayment
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}16'] = -100000
    
    # Dividends Paid
    dividend_values = [-50000, -55000, -60500, -66550, -73205]
    
    for j, (year, dividend) in enumerate(zip(years, dividend_values), 2):
        col = get_column_letter(j)
        ws[f'{col}17'] = dividend
    
    # Cash Flow from Financing Activities
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}18'] = f"=SUM({col}16:{col}17)"
        ws[f'{col}18'].font = Font(bold=True)
    
    # Net Change in Cash
    ws['A20'] = "Net Change in Cash"
    ws['A20'].font = Font(bold=True)
    
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}20'] = f"={col}9+{col}13+{col}18"
        ws[f'{col}20'].font = Font(bold=True)
    
    # Beginning Cash Balance
    ws['A21'] = "Beginning Cash Balance"
    ws['B21'] = 500000  # Starting cash for 2025
    
    for j, year in enumerate(years[1:], 3):
        col = get_column_letter(j)
        prev_col = get_column_letter(j-1)
        ws[f'{col}21'] = f"={prev_col}22"  # Equal to previous year ending balance
    
    # Ending Cash Balance
    ws['A22'] = "Ending Cash Balance"
    ws['A22'].font = Font(bold=True)
    
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}22'] = f"={col}21+{col}20"
        ws[f'{col}22'].font = Font(bold=True)
    
    # Now update the Balance_Sheet Cash formulas to link to this Cash_Flow sheet
    balance_sheet = wb["Balance_Sheet"]
    
    # First year (2025) is already set with the initial value
    # Update formulas for future years
    for j, year in enumerate(years[1:], 3):
        col = get_column_letter(j)
        balance_sheet[f'{col}6'] = f"=Cash_Flow!{col}22"
    
    # Format all monetary values with thousands separator
    for row in range(6, 23):
        for col in range(2, 7):  # Columns B to F
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.number_format = '#,##0'
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 7):  # Columns B to F
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add background colors to key rows
    op_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')       # Light green
    inv_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')      # Light blue
    fin_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')      # Light orange
    change_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # Light yellow
    
    # Apply fills
    for col in range(1, 7):  # Columns A to F
        col_letter = get_column_letter(col)
        ws[f'{col_letter}9'].fill = op_fill       # Cash Flow from Operating Activities
        ws[f'{col_letter}13'].fill = inv_fill     # Cash Flow from Investing Activities
        ws[f'{col_letter}18'].fill = fin_fill     # Cash Flow from Financing Activities
        ws[f'{col_letter}20'].fill = change_fill  # Net Change in Cash
        ws[f'{col_letter}22'].fill = change_fill  # Ending Cash Balance
    
    print("Cash_Flow sheet created successfully")