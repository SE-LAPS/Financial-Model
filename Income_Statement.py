from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series

def create_income_statement(wb):
    """Create and format the Income_Statement sheet."""
    print("Creating Income_Statement sheet...")
    
    # Get the Income_Statement sheet
    ws = wb["Income_Statement"]
    
    # Set the title
    ws['A1'] = "INCOME STATEMENT"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Column headers
    ws['A3'] = "Line Item"
    
    # Year headers
    years = ["2025", "2026", "2027", "2028", "2029"]
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)  # Start from column B
        ws[f'{col}3'] = year
        ws[f'{col}3'].font = Font(bold=True)
    
    # Make header row bold
    ws['A3'].font = Font(bold=True)
    
    # Income Statement line items
    line_items = [
        "Revenue",
        "Cost of Revenue",
        "Gross Profit",
        "Operating Expenses:",
        "   Sales & Marketing",
        "   Research & Development",
        "   General & Administrative",
        "   Depreciation & Amortization",
        "Total Operating Expenses",
        "Operating Income",
        "EBITDA",
        "Interest Expense",
        "Other Income/(Expense)",
        "Earnings Before Tax",
        "Tax Expense",
        "Net Income"
    ]
    
    for i, item in enumerate(line_items, 5):
        row = i
        ws[f'A{row}'] = item
        
        # Make key line items bold
        if item in ["Revenue", "Gross Profit", "Total Operating Expenses", "Operating Income", "EBITDA", "Earnings Before Tax", "Net Income"]:
            ws[f'A{row}'].font = Font(bold=True)
    
    # Revenue (linked to Revenue_Forecast)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}5'] = f"=Revenue_Forecast!{col}9"
        ws[f'{col}5'].font = Font(bold=True)
    
    # Cost of Revenue (75% gross margin target)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}6'] = f"={col}5*0.25"  # 25% of revenue
    
    # Gross Profit
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}7'] = f"={col}5-{col}6"
        ws[f'{col}7'].font = Font(bold=True)
    
    # Operating Expenses
    # Sales & Marketing (30% of revenue)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}9'] = f"={col}5*0.30"
    
    # R&D (15% of revenue)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}10'] = f"={col}5*0.15"
    
    # G&A (10% of revenue)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}11'] = f"={col}5*0.10"
    
    # D&A (5% of revenue)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}12'] = f"={col}5*0.05"
    
    # Total Operating Expenses
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}13'] = f"=SUM({col}9:{col}12)"
        ws[f'{col}13'].font = Font(bold=True)
    
    # Operating Income
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}14'] = f"={col}7-{col}13"
        ws[f'{col}14'].font = Font(bold=True)
    
    # EBITDA
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}15'] = f"={col}14+{col}12"  # Operating Income + D&A
        ws[f'{col}15'].font = Font(bold=True)
    
    # Interest Expense (decreasing as company grows)
    interest_values = [2000000, 1800000, 1500000, 1200000, 1000000]
    for j, value in enumerate(interest_values, 2):
        col = get_column_letter(j)
        ws[f'{col}16'] = value
    
    # Other Income/(Expense) (investment income, 0.5% of revenue)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}17'] = f"={col}5*0.005"
    
    # Earnings Before Tax
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}18'] = f"={col}14-{col}16+{col}17"
        ws[f'{col}18'].font = Font(bold=True)
    
    # Tax Expense (25% effective tax rate)
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}19'] = f"={col}18*0.25"
    
    # Net Income
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}20'] = f"={col}18-{col}19"
        ws[f'{col}20'].font = Font(bold=True)
    
    # Format all numbers
    for row in range(5, 21):
        for col in range(2, 7):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.number_format = '#,##0'
    
    # Add margin calculations
    margin_items = [
        ("Gross Margin %", "=B7/B5"),
        ("Operating Margin %", "=B14/B5"),
        ("EBITDA Margin %", "=B15/B5"),
        ("Net Margin %", "=B20/B5")
    ]
    
    for i, (item, formula) in enumerate(margin_items):
        row = i + 22
        ws[f'A{row}'] = item
        ws[f'A{row}'].font = Font(bold=True)
        
        for j, year in enumerate(years, 2):
            col = get_column_letter(j)
            base_formula = formula.replace('B', col)
            ws[f'{col}{row}'] = base_formula
            ws[f'{col}{row}'].number_format = '0.00%'
    
    # Add a chart for key margins
    chart = LineChart()
    chart.title = "Key Margin Trends"
    chart.style = 10
    chart.x_axis.title = "Year"
    chart.y_axis.title = "Margin %"
    
    # Add data series
    cats = Reference(ws, min_col=2, min_row=3, max_col=6, max_row=3)
    for row in range(22, 26):
        data = Reference(ws, min_col=2, min_row=row, max_col=6, max_row=row)
        series = Series(data, title=ws[f'A{row}'].value)
        chart.append(series)
    
    chart.set_categories(cats)
    ws.add_chart(chart, "A28")
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add background colors
    header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    subtotal_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Apply fills
    ws['A1'].fill = header_fill
    ws['A3'].fill = header_fill
    for col in range(2, 7):
        ws[f'{get_column_letter(col)}3'].fill = header_fill
    
    # Apply subtotal fills
    for row in [7, 13, 14, 15, 18, 20]:  # Gross Profit, Total OpEx, Operating Income, EBITDA, EBT, Net Income
        ws[f'A{row}'].fill = subtotal_fill
        for col in range(2, 7):
            ws[f'{get_column_letter(col)}{row}'].fill = subtotal_fill
    
    print("Income_Statement sheet created successfully")