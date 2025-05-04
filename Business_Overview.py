from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_business_overview(wb):
    """Create and format the Business_Overview sheet."""
    print("Creating Business_Overview sheet...")
    
    # Get the Business_Overview sheet
    ws = wb["Business_Overview"]
    
    # Set the title
    ws['A1'] = "COMPANY OVERVIEW"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Company information
    ws['A3'] = "Company Name:"
    ws['B3'] = "TechVision Solutions Inc."
    
    ws['A4'] = "Industry:"
    ws['B4'] = "Enterprise Software & Cloud Services"
    
    ws['A5'] = "Business Model:"
    ws['B5'] = "B2B SaaS provider offering enterprise software solutions with a subscription-based revenue model"
    
    # Company Description
    ws['A7'] = "Company Description"
    ws['A7'].font = Font(bold=True)
    ws['A8'] = "TechVision Solutions Inc. is a leading provider of enterprise software solutions, specializing in cloud-based business intelligence, data analytics, and process automation tools. Founded in 2020, the company has rapidly grown to serve over 500 enterprise clients across multiple industries."
    
    # Products and Services section
    ws['A10'] = "Products and Services"
    ws['A10'].font = Font(bold=True)
    
    products = [
        ("DataInsight Pro", "Advanced business intelligence and analytics platform", 40),
        ("CloudFlow", "Cloud-based workflow automation solution", 30),
        ("SecureConnect", "Enterprise security and integration platform", 20),
        ("AI Assistant", "AI-powered business process automation tool", 10)
    ]
    
    for i, (product, desc, revenue) in enumerate(products):
        row = i + 12
        ws[f'A{row}'] = product
        ws[f'B{row}'] = desc
        ws[f'C{row}'] = f"{revenue}%"
        ws[f'A{row}'].font = Font(bold=True)
    
    # Market Analysis
    ws['A17'] = "Market Analysis"
    ws['A17'].font = Font(bold=True)
    
    market_points = [
        "Total Addressable Market (TAM): $50 billion",
        "Serviceable Addressable Market (SAM): $20 billion",
        "Serviceable Obtainable Market (SOM): $2 billion",
        "Expected CAGR: 15% (2025-2029)",
        "Key Growth Drivers: Digital transformation, AI adoption, cloud migration"
    ]
    
    for i, point in enumerate(market_points):
        ws[f'A{19+i}'] = point
    
    # Competitive Advantages
    ws['A26'] = "Competitive Advantages"
    ws['A26'].font = Font(bold=True)
    
    advantages = [
        "Proprietary AI/ML technology",
        "Strong IP portfolio with 15 patents",
        "99.9% platform uptime",
        "24/7 enterprise support",
        "ISO 27001 certified security"
    ]
    
    for i, advantage in enumerate(advantages):
        ws[f'A{28+i}'] = "• " + advantage
    
    # Growth Strategy
    ws['A35'] = "Growth Strategy"
    ws['A35'].font = Font(bold=True)
    
    strategies = [
        "Geographic expansion into APAC region",
        "New product development in AI/ML space",
        "Strategic acquisitions in complementary technologies",
        "Channel partner program expansion",
        "Investment in R&D (15% of revenue)"
    ]
    
    for i, strategy in enumerate(strategies):
        ws[f'A{37+i}'] = "• " + strategy
    
    # Financial Highlights
    ws['A44'] = "Financial Highlights (2024)"
    ws['A44'].font = Font(bold=True)
    
    highlights = [
        ("Annual Revenue", "$100 million"),
        ("Gross Margin", "75%"),
        ("EBITDA Margin", "25%"),
        ("ARR Growth", "35%"),
        ("Customer Retention", "95%")
    ]
    
    for i, (metric, value) in enumerate(highlights):
        ws[f'A{46+i}'] = metric
        ws[f'B{46+i}'] = value
    
    # Formatting
    # Wrap text in description cells
    for row in range(8, 50):
        for col in ['A', 'B']:
            cell = ws[f'{col}{row}']
            cell.alignment = Alignment(wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    
    # Add background colors
    header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    section_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Apply fills to main header and section headers
    ws['A1'].fill = header_fill
    for row in [7, 10, 17, 26, 35, 44]:
        ws[f'A{row}'].fill = section_fill
    
    print("Business_Overview sheet created successfully")