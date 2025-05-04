from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

def create_dashboard(wb):
    """Create and format the Dashboard sheet."""
    print("Creating Dashboard sheet...")
    
    # Get the Dashboard sheet
    ws = wb["Dashboard"]
    
    # Set the title
    ws['A1'] = "FINANCIAL MODEL DASHBOARD"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Key Financial Metrics section
    ws['A3'] = "Key Financial Metrics"
    ws['A3'].font = Font(bold=True)
    
    # Metrics labels and formulas
    metrics = [
        ("Revenue (USD)", "=Revenue_Forecast!B9"),
        ("Gross Profit (USD)", "=Income_Statement!B7"),
        ("EBITDA (USD)", "=Income_Statement!B10"),
        ("Net Income (USD)", "=Income_Statement!B14"),
        ("Cash Balance (USD)", "=Balance_Sheet!B6"),
        ("NPV (USD)", "=Capital_Budgeting!B18"),
        ("IRR", "=Capital_Budgeting!B22"),
        ("Payback Period (Years)", "=Capital_Budgeting!B26")
    ]
    
    for i, (metric, formula) in enumerate(metrics, 5):
        row = i
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = formula
        
        # Format numbers
        if "IRR" in metric:
            ws[f'B{row}'].number_format = '0.00%'
        elif "Period" in metric:
            ws[f'B{row}'].number_format = '0.00'
        else:
            ws[f'B{row}'].number_format = '#,##0'
    
    # Financial Ratios section
    ws['D3'] = "Financial Ratios"
    ws['D3'].font = Font(bold=True)
    
    ratios = [
        ("Gross Margin", "=Income_Statement!B7/Income_Statement!B5"),
        ("EBITDA Margin", "=Income_Statement!B10/Income_Statement!B5"),
        ("Net Profit Margin", "=Income_Statement!B14/Income_Statement!B5"),
        ("ROE", "=Income_Statement!B14/Balance_Sheet!B20"),
        ("Current Ratio", "=Balance_Sheet!B8/Balance_Sheet!B15")
    ]
    
    for i, (ratio, formula) in enumerate(ratios, 5):
        row = i
        ws[f'D{row}'] = ratio
        ws[f'E{row}'] = formula
        ws[f'E{row}'].number_format = '0.00%' if ratio != "Current Ratio" else '0.00'
    
    # Revenue Forecast Chart
    ws['A15'] = "Revenue Forecast"
    ws['A15'].font = Font(bold=True)
    
    # Years for the chart
    years = ["2025", "2026", "2027", "2028", "2029"]
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)
        ws[f'{col}15'] = year
    
    # Revenue values
    for i in range(5):
        col = get_column_letter(i + 2)
        ws[f'{col}16'] = f"=Revenue_Forecast!{col}9"
        ws[f'{col}16'].number_format = '#,##0'
    
    chart1 = BarChart()
    chart1.title = "Revenue Forecast 2025-2029"
    chart1.style = 10
    chart1.x_axis.title = "Year"
    chart1.y_axis.title = "Revenue"
    
    data = Reference(ws, min_col=2, min_row=16, max_col=6, max_row=16)
    cats = Reference(ws, min_col=2, min_row=15, max_col=6, max_row=15)
    chart1.add_data(data)
    chart1.set_categories(cats)
    
    ws.add_chart(chart1, "A20")
    
    # Income Statement Trends
    ws['A40'] = "Income Statement Trends"
    ws['A40'].font = Font(bold=True)
    
    # Copy data for visualization
    metrics = ["Revenue", "Gross Profit", "EBITDA", "Net Income"]
    for i, metric in enumerate(metrics):
        ws[f'A{41+i}'] = metric
    
    for j, year in enumerate(years):
        col = get_column_letter(j + 2)
        ws[f'{col}40'] = year
        
        # Get values from Income Statement
        ws[f'{col}41'] = f"=Revenue_Forecast!{col}9"  # Revenue
        ws[f'{col}42'] = f"=Income_Statement!{col}7"  # Gross Profit
        ws[f'{col}43'] = f"=Income_Statement!{col}10"  # EBITDA
        ws[f'{col}44'] = f"=Income_Statement!{col}14"  # Net Income
        
        for row in range(41, 45):
            ws[f'{col}{row}'].number_format = '#,##0'
    
    chart2 = LineChart()
    chart2.title = "Income Statement Trends"
    chart2.style = 10
    chart2.x_axis.title = "Year"
    chart2.y_axis.title = "Amount"
    
    data = Reference(ws, min_col=2, min_row=41, max_col=6, max_row=44)
    cats = Reference(ws, min_col=2, min_row=40, max_col=6, max_row=40)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    
    ws.add_chart(chart2, "G20")
    
    # Cost Structure (Pie Chart)
    ws['A50'] = "Cost Structure (Latest Year)"
    ws['A50'].font = Font(bold=True)
    
    cost_items = [
        ("COGS", "=Income_Statement!B6"),
        ("Operating Expenses", "=Income_Statement!B8"),
        ("D&A", "=Income_Statement!B9"),
        ("Interest", "=Income_Statement!B11"),
        ("Tax", "=Income_Statement!B13")
    ]
    
    for i, (item, formula) in enumerate(cost_items):
        ws[f'A{51+i}'] = item
        ws[f'B{51+i}'] = formula
        ws[f'B{51+i}'].number_format = '#,##0'
    
    chart3 = PieChart()
    chart3.title = "Cost Structure"
    
    labels = Reference(ws, min_col=1, min_row=51, max_row=55)
    data = Reference(ws, min_col=2, min_row=51, max_row=55)
    chart3.add_data(data)
    chart3.set_categories(labels)
    
    ws.add_chart(chart3, "A65")
    
    # NPV Sensitivity (from Sensitivity Analysis)
    ws['G50'] = "NPV Sensitivity"
    ws['G50'].font = Font(bold=True)
    
    # Copy sensitivity data
    for i in range(5):
        row = i + 51
        ws[f'G{row}'] = f"=Sensitivity_Analysis!A{i+9}"  # Discount rates
        ws[f'H{row}'] = f"=Sensitivity_Analysis!B{i+9}"  # NPV values
        ws[f'G{row}'].number_format = '0.00%'
        ws[f'H{row}'].number_format = '#,##0'
    
    chart4 = ScatterChart()
    chart4.title = "NPV Sensitivity to Discount Rate"
    chart4.style = 10
    chart4.x_axis.title = "Discount Rate"
    chart4.y_axis.title = "NPV"
    
    xvalues = Reference(ws, min_col=7, min_row=51, max_row=55)
    yvalues = Reference(ws, min_col=8, min_row=51, max_row=55)
    series = Series(yvalues, xvalues, title="NPV")
    chart4.series.append(series)
    
    ws.add_chart(chart4, "G65")
    
    # Add conditional formatting for financial metrics
    green_to_red = ColorScaleRule(start_type='max', start_color='63BE7B',
                               mid_type='percentile', mid_value=50, mid_color='FFEB84',
                               end_type='min', end_color='F8696B')
    
    ws.conditional_formatting.add('B5:B12', green_to_red)  # Key metrics
    ws.conditional_formatting.add('E5:E9', green_to_red)   # Financial ratios
    
    # Set column widths
    for col in ['A', 'D']:
        ws.column_dimensions[col].width = 25
    for col in ['B', 'C', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15
    
    print("Dashboard sheet created successfully")