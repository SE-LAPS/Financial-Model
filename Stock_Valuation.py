from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def create_stock_valuation(wb):
    """Create and format the Stock_Valuation sheet."""
    print("Creating Stock_Valuation sheet...")
    
    # Get the Stock_Valuation sheet
    ws = wb["Stock_Valuation"]
    
    # Set the title
    ws['A1'] = "STOCK VALUATION MODEL"
    ws['A1'].font = Font(bold=True, size=14)
    
    # DCF Valuation section
    ws['A3'] = "Discounted Cash Flow (DCF) Valuation"
    ws['A3'].font = Font(bold=True, size=12)
    
    # Basic inputs
    ws['A5'] = "WACC (Discount Rate)"
    ws['B5'] = "=Assumptions!B22"
    ws['B5'].number_format = '0.00%'
    
    ws['A6'] = "Long-term Growth Rate"
    ws['B6'] = 0.025
    ws['B6'].number_format = '0.00%'
    
    ws['A7'] = "Shares Outstanding"
    ws['B7'] = 1000000
    ws['B7'].number_format = '#,##0'
    
    # Free Cash Flow section
    ws['A9'] = "Free Cash Flow"
    ws['A9'].font = Font(bold=True)
    
    # Column headers for years and terminal value
    years = ["2025", "2026", "2027", "2028", "2029", "Terminal"]
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)  # Start from column B
        ws[f'{col}9'] = year
        ws[f'{col}9'].font = Font(bold=True)
    
    # Free Cash Flow line items
    fcf_items = [
        "EBIT",
        "Tax Rate",
        "EBIT*(1-Tax Rate)",
        "Plus: Depreciation & Amortization",
        "Less: Capital Expenditures",
        "Less: Change in Working Capital",
        "Free Cash Flow"
    ]
    
    for i, item in enumerate(fcf_items, 10):
        row = i
        ws[f'A{row}'] = item
        if item == "Free Cash Flow":
            ws[f'A{row}'].font = Font(bold=True)
    
    # EBIT formulas
    for j, year in enumerate(years[:5], 2):  # 2025-2029
        col = get_column_letter(j)
        ws[f'{col}10'] = f"=Income_Statement!{col}9"  # Operating Income
    
    # Terminal year EBIT
    ws['G10'] = "=F10*(1+B6)"  # Terminal year EBIT = 2029 EBIT * (1 + growth rate)
    
    # Tax Rate formulas
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}11'] = "=Assumptions!B6"
        ws[f'{col}11'].number_format = '0.00%'
    
    # EBIT*(1-Tax Rate) formulas
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}12'] = f"={col}10*(1-{col}11)"
    
    # Depreciation & Amortization formulas
    for j, year in enumerate(years[:5], 2):  # 2025-2029
        col = get_column_letter(j)
        ws[f'{col}13'] = f"=OPEX_Budget!{col}7"
    
    # Terminal year D&A
    ws['G13'] = "=F13*(1+B6)"  # Terminal year D&A = 2029 D&A * (1 + growth rate)
    
    # Capital Expenditures formulas
    for j, year in enumerate(years[:5], 2):  # 2025-2029
        col = get_column_letter(j)
        ws[f'{col}14'] = f"=-Cash_Flow!{col}12"  # Negative of Cash_Flow CapEx (which is already negative)
    
    # Terminal year CapEx (assumed equal to D&A for terminal value)
    ws['G14'] = "=G13"  # Terminal year CapEx = Terminal year D&A
    
    # Change in Working Capital formulas
    for j, year in enumerate(years[:5], 2):  # 2025-2029
        col = get_column_letter(j)
        ws[f'{col}15'] = f"=-Cash_Flow!{col}8"  # Negative of Cash_Flow WC change
    
    # Terminal year Change in WC
    ws['G15'] = "=F15*(1+B6)"  # Terminal year WC = 2029 WC * (1 + growth rate)
    
    # Free Cash Flow formulas
    for j, year in enumerate(years, 2):
        col = get_column_letter(j)
        ws[f'{col}16'] = f"={col}12+{col}13+{col}14+{col}15"
        ws[f'{col}16'].font = Font(bold=True)
    
    # Terminal Value calculation
    ws['A18'] = "Terminal Value"
    ws['B18'] = "=G16/(B5-B6)"  # Terminal Value using Gordon Growth Model
    
    # Discount Factor calculations
    ws['A19'] = "Discount Factor"
    
    for j, year in enumerate(years[:5], 2):  # 2025-2029
        col = get_column_letter(j)
        ws[f'{col}19'] = f"=1/(1+B5)^{j-1}"  # 1/(1+WACC)^year
        ws[f'{col}19'].number_format = '0.000'
    
    # Terminal year discount factor (same as 2029)
    ws['G19'] = "=F19"
    ws['G19'].number_format = '0.000'
    
    # Present Value of FCF
    ws['A20'] = "Present Value of FCF"
    
    for j, year in enumerate(years[:5], 2):  # 2025-2029
        col = get_column_letter(j)
        ws[f'{col}20'] = f"={col}16*{col}19"
    
    # Present Value of Terminal Value
    ws['G20'] = "=B18*G19"
    
    # Sum of PV of FCF
    ws['A22'] = "Sum of PV of FCF"
    ws['A22'].font = Font(bold=True)
    ws['B22'] = "=SUM(B20:F20)"
    ws['B22'].font = Font(bold=True)
    
    # PV of Terminal Value
    ws['A23'] = "PV of Terminal Value"
    ws['B23'] = "=G20"
    
    # Enterprise Value
    ws['A24'] = "Enterprise Value"
    ws['A24'].font = Font(bold=True)
    ws['B24'] = "=B22+B23"
    ws['B24'].font = Font(bold=True)
    
    # Less: Net Debt
    ws['A26'] = "Less: Net Debt"
    ws['B26'] = "=Balance_Sheet!B14-Balance_Sheet!B6"  # Debt - Cash
    
    # Equity Value
    ws['A27'] = "Equity Value"
    ws['A27'].font = Font(bold=True)
    ws['B27'] = "=B24-B26"
    ws['B27'].font = Font(bold=True)
    
    # Share Price
    ws['A28'] = "Share Price"
    ws['A28'].font = Font(bold=True)
    ws['B28'] = "=B27/B7"
    ws['B28'].number_format = '$#,##0.00'
    ws['B28'].font = Font(bold=True)
    
    # Comparable Company Valuation section
    ws['A30'] = "Comparable Company Valuation"
    ws['A30'].font = Font(bold=True, size=12)
    
    # EV/EBITDA Multiple Valuation
    ws['A32'] = "EV/EBITDA Multiple"
    ws['B32'] = 8
    ws['B32'].number_format = '0.0'
    
    ws['A33'] = "EBITDA (Last Year)"
    ws['B33'] = "=Income_Statement!F10"  # 2029 EBITDA
    
    ws['A34'] = "Enterprise Value"
    ws['B34'] = "=B32*B33"
    
    ws['A35'] = "Less: Net Debt"
    ws['B35'] = "=B26"
    
    ws['A36'] = "Equity Value"
    ws['A36'].font = Font(bold=True)
    ws['B36'] = "=B34-B35"
    ws['B36'].font = Font(bold=True)
    
    ws['A37'] = "Share Price"
    ws['A37'].font = Font(bold=True)
    ws['B37'] = "=B36/B7"
    ws['B37'].number_format = '$#,##0.00'
    ws['B37'].font = Font(bold=True)
    
    # P/E Multiple Valuation
    ws['A39'] = "P/E Multiple"
    ws['B39'] = 15
    ws['B39'].number_format = '0.0'
    
    ws['A40'] = "EPS (Last Year)"
    ws['B40'] = "=Income_Statement!F14/B7"  # 2029 Net Income / Shares Outstanding
    ws['B40'].number_format = '$#,##0.00'
    
    ws['A41'] = "Share Price"
    ws['A41'].font = Font(bold=True)
    ws['B41'] = "=B39*B40"
    ws['B41'].number_format = '$#,##0.00'
    ws['B41'].font = Font(bold=True)
    
    # Format all monetary values with thousands separator
    for row in range(10, 17):  # FCF table
        for col in range(2, 8):  # Columns B to G
            if row != 11:  # Skip tax rate row
                cell = ws[f'{get_column_letter(col)}{row}']
                cell.number_format = '#,##0'
    
    # Format other monetary values
    for row in [18, 20, 22, 23, 24, 26, 27, 33, 34, 35, 36]:
        ws[f'B{row}'].number_format = '#,##0'
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 8):  # Columns B to G
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add background colors to key rows
    dcf_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')       # Light blue
    ev_ebitda_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Light green
    pe_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')        # Light orange
    
    # Apply fills to key rows
    for col in range(1, 3):  # Columns A to B for summary rows
        col_letter = get_column_letter(col)
        ws[f'{col_letter}24'].fill = dcf_fill      # Enterprise Value
        ws[f'{col_letter}27'].fill = dcf_fill      # Equity Value
        ws[f'{col_letter}28'].fill = dcf_fill      # Share Price (DCF)
        ws[f'{col_letter}36'].fill = ev_ebitda_fill # Equity Value (EV/EBITDA)
        ws[f'{col_letter}37'].fill = ev_ebitda_fill # Share Price (EV/EBITDA)
        ws[f'{col_letter}41'].fill = pe_fill        # Share Price (P/E)
    
    # Apply fills to FCF row
    for col in range(1, 8):  # Columns A to G
        col_letter = get_column_letter(col)
        ws[f'{col_letter}16'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow
    
    print("Stock_Valuation sheet created successfully")