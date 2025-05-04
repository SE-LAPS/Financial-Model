from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def create_opex_budget(wb):
    """Create and format the OPEX_Budget sheet."""
    print("Creating OPEX_Budget sheet...")
    
    # Get the OPEX_Budget sheet
    ws = wb["OPEX_Budget"]
    
    # Set the title
    ws['A1'] = "OPERATING EXPENSES BUDGET"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Column headers
    ws['A3'] = "Category"
    
    # Year headers
    years = ["2025", "2026", "2027", "2028", "2029"]
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)  # Start from column B
        ws[f'{col}3'] = year
        ws[f'{col}3'].font = Font(bold=True)
    
    # Make header row bold
    ws['A3'].font = Font(bold=True)
    
    # Operating expenses
    expenses = ["SG&A Expenses", "R&D Expenses", "Depreciation & Amortization"]
    
    # SG&A Expenses
    ws['A5'] = expenses[0]
    for j, year in enumerate(years, 2):  # Start from column B
        col = get_column_letter(j)
        ws[f'{col}5'] = f"=Revenue_Forecast!{col}9*Assumptions!B18"
    
    # R&D Expenses
    ws['A6'] = expenses[1]
    for j, year in enumerate(years, 2):  # Start from column B
        col = get_column_letter(j)
        ws[f'{col}6'] = f"=Revenue_Forecast!{col}9*Assumptions!B19"
    
    # Depreciation & Amortization (fixed values)
    ws['A7'] = expenses[2]
    depreciation_values = [100000, 105000, 110250, 115763, 121551]
    for j, (year, value) in enumerate(zip(years, depreciation_values), 2):
        col = get_column_letter(j)
        ws[f'{col}7'] = value
    
    # Total Operating Expenses row
    ws['A9'] = "Total Operating Expenses"
    ws['A9'].font = Font(bold=True)
    
    # Sum formulas for total operating expenses
    for i, year in enumerate(years, 2):  # Start from column B
        col = get_column_letter(i)
        ws[f'{col}9'] = f"=SUM({col}5:{col}7)"
        ws[f'{col}9'].font = Font(bold=True)
    
    # Format all expense values with thousands separator
    for row in range(5, 10):
        for col in range(2, 7):  # Columns B to F
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.number_format = '#,##0'
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    for col in range(2, 7):  # Columns B to F
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add a light orange background to the total row
    total_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    for col in range(1, 7):  # Columns A to F
        ws[f'{get_column_letter(col)}9'].fill = total_fill
    
    print("OPEX_Budget sheet created successfully")